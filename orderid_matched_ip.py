import openpyxl
import pyautogui
import time
import pyperclip

def search_excel_column(filename, filenameOut, column_letter, search_box_position,search_button_position):
    """
    filename:读取的xlsx文件
    filenameOut:写入的xlsx文件
    column_letter:读取filename时要读取的列（如A列）
    search_box_position:坐标(搜索框的坐标)
    search_button_position:坐标（搜索按钮的坐标）
    """


    #打开 Excel 文件
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # 找到指定列并逐行读取数据
    column = sheet[column_letter]
    current_row = 1
    for cell in column:
        search_term = str(cell.value)  # 将单元格的值转换为字符串
        if search_term.strip():  # 确保单元格不为空
            # 清除搜索框中已有内容
            pyautogui.click(search_box_position)
            pyautogui.hotkey('ctrl', 'a')  # 选择所有内容
            pyautogui.press('delete')  # 删除已有内容

            # 在搜索框中输入数据并搜索
            pyautogui.typewrite(search_term)
            pyautogui.press('enter')
            time.sleep(0.1)
            pyautogui.click(search_button_position)
            time.sleep(2)  # 等待页面加载

            wbOut = openpyxl.load_workbook(filenameOut)
            sheetOut = wbOut.active
            double_click_position = (1769, 748)  # 请替换为实际的坐标
            pyautogui.doubleClick(double_click_position)
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'c')
            selected_text = pyperclip.paste()  # 需要根据您的实际情况调整按键间隔

            # 将文本写入 Excel 的 B 列，每次延后一行
            sheet.cell(row=current_row, column=2, value=selected_text)
            current_row += 1

            # 保存 Excel 文件
            wb.save(filenameOut)

            # 延迟一段时间以防止重复写入
            time.sleep(1)  # 根据需要调整延迟时间



# 指定 Excel 文件名、要读取的列（例如'A'）、搜索框位置（以屏幕坐标表示）
filename = 'test1.xlsx'
filenameOut = 'test1.xlsx'
column_letter = 'A'
search_box_position = (949, 334)  # 替换为实际的搜索框位置
search_button_position = (1763, 335)

# 调用函数进行搜索
time.sleep(5)
search_excel_column(filename, filenameOut,column_letter, search_box_position, search_button_position)
